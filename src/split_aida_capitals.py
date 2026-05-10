"""
split_aida_capitals.py
======================
Splits an AIDA / Orbis exported Excel file into per-capital files
following a 4-Capital strategic valuation model:

    1. Financial Capital
    2. Technological Capital
    3. Human & Organisational Capital
    4. Relational Capital

Context columns (company identifiers, sector codes, peer group) are
included in EVERY output file and EVERY sheet of the combined workbook.

Outputs (saved to ./capital_split_outputs/):
    context_columns.xlsx
    financial_capital.xlsx
    technological_capital.xlsx
    human_organisational_capital.xlsx
    relational_capital.xlsx
    all_capitals_clean_split.xlsx   (multi-sheet workbook)

Run:
    pip install pandas openpyxl xlrd
    python split_aida_capitals.py
"""

from __future__ import annotations

import os
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIG  -- edit these if you need to
# =============================================================================
INPUT_FILE = "aida_export.xlsx"     # path to your consolidated AIDA export
OUTPUT_DIR = "capital_split_outputs"
SHEET_NAME = 0                       # 0 = first sheet, or pass a sheet name


# =============================================================================
# COLUMN GROUP DEFINITIONS
# =============================================================================
# Context columns appear in every output file / every sheet.
context_columns = [
    "Company name",
    "Province",
    "Accounting closing date Last avail. yr",
    "Tax code number",
    "CCIAA number",
    "Primary business line",
    "Main activity",
    "ATECO 2007 code",
    "ATECO 2007 description",
    "NACE Rev. 2",
    "NACE Rev. 2 description",
    "Peer Group Name",
    "Peer Group Description",
    "Peer Group Size",
    "Main products and services",
    "Size estimate",
]

# 1. FINANCIAL CAPITAL: profitability, growth, liquidity, debt, cash quality
financial_columns = [
    "Revenues from sales and services th EUR Last avail. yr",
    "Revenues from sales and services th EUR 2024",
    "Revenues from sales and services th EUR 2023",
    "Revenues from sales and services th EUR 2022",
    "EBITDA th EUR Last avail. yr",
    "EBITDA th EUR 2024",
    "EBITDA th EUR 2023",
    "EBITDA th EUR 2022",
    "Profit (loss) th EUR Last avail. yr",
    "Profit (loss) th EUR 2024",
    "Profit (loss) th EUR 2023",
    "Profit (loss) th EUR 2022",
    "EBITDA/Vendite % Last avail. yr",
    "EBITDA/Vendite % 2024",
    "EBITDA/Vendite % 2023",
    "EBITDA/Vendite % 2022",
    "Total shareholder's funds th EUR Last avail. yr",
    "Total shareholder's funds th EUR 2024",
    "Total shareholder's funds th EUR 2023",
    "Total shareholder's funds th EUR 2022",
    "Total assets th EUR Last avail. yr",
    "Total assets th EUR 2024",
    "Total assets th EUR 2023",
    "Total assets th EUR 2022",
    "Net financial position th EUR Last avail. yr",
    "Net financial position th EUR 2024",
    "Net financial position th EUR 2023",
    "Net financial position th EUR 2022",
    "Debt/EBITDA ratio % Last avail. yr",
    "Debt/EBITDA ratio % 2024",
    "Debt/EBITDA ratio % 2023",
    "Debt/EBITDA ratio % 2022",
    "Current ratio Last avail. yr",
    "Current ratio 2024",
    "Current ratio 2023",
    "Current ratio 2022",
    "Cash flow from operating activities (A) th EUR Last avail. yr",
    "Cash flow from operating activities (A) th EUR 2024",
    "Cash flow from operating activities (A) th EUR 2023",
    "Cash flow from operating activities (A) th EUR 2022",
]

# 2. TECHNOLOGICAL CAPITAL: fixed/tangible/intangible assets, R&D, IP, investment
technological_columns = [
    "TOTAL FIXED ASSETS th EUR Last avail. yr",
    "TOTAL FIXED ASSETS th EUR 2024",
    "TOTAL FIXED ASSETS th EUR 2023",
    "TOTAL FIXED ASSETS th EUR 2022",
    "TOTAL INTANGIBLE FIXED ASSETS th EUR Last avail. yr",
    "TOTAL INTANGIBLE FIXED ASSETS th EUR 2024",
    "TOTAL INTANGIBLE FIXED ASSETS th EUR 2023",
    "TOTAL INTANGIBLE FIXED ASSETS th EUR 2022",
    "Plant and machinery th EUR Last avail. yr",
    "Plant and machinery th EUR 2024",
    "Plant and machinery th EUR 2023",
    "Plant and machinery th EUR 2022",
    "Indust. and commercial equipment th EUR Last avail. yr",
    "Indust. and commercial equipment th EUR 2024",
    "Indust. and commercial equipment th EUR 2023",
    "Indust. and commercial equipment th EUR 2022",
    "TOTAL TANGIBLE FIXED ASSETS th EUR Last avail. yr",
    "TOTAL TANGIBLE FIXED ASSETS th EUR 2024",
    "TOTAL TANGIBLE FIXED ASSETS th EUR 2023",
    "TOTAL TANGIBLE FIXED ASSETS th EUR 2022",
    "Research and dev. exp. th EUR Last avail. yr",
    "Research and dev. exp. th EUR 2024",
    "Research and dev. exp. th EUR 2023",
    "Research and dev. exp. th EUR 2022",
    "Ind. patents and intellect. property rights th EUR Last avail. yr",
    "Ind. patents and intellect. property rights th EUR 2024",
    "Ind. patents and intellect. property rights th EUR 2023",
    "Ind. patents and intellect. property rights th EUR 2022",
    "Concessions, licenses, trademarks and similar rights th EUR Last avail. yr",
    "Concessions, licenses, trademarks and similar rights th EUR 2024",
    "Concessions, licenses, trademarks and similar rights th EUR 2023",
    "Concessions, licenses, trademarks and similar rights th EUR 2022",
    "Tangible assets: (investment) th EUR Last avail. yr",
    "Tangible assets: (investment) th EUR 2024",
    "Tangible assets: (investment) th EUR 2023",
    "Tangible assets: (investment) th EUR 2022",
    "Intangible assets: (investment) th EUR Last avail. yr",
    "Intangible assets: (investment) th EUR 2024",
    "Intangible assets: (investment) th EUR 2023",
    "Intangible assets: (investment) th EUR 2022",
]

# 3. HUMAN & ORGANISATIONAL CAPITAL: workforce, productivity, governance, control
human_organisational_columns = [
    "Number of employees Last avail. yr",
    "Employees Last avail. yr",
    "Employees 2024",
    "Employees 2023",
    "Employees 2022",
    "Total personnel costs th EUR Last avail. yr",
    "Total personnel costs th EUR 2024",
    "Total personnel costs th EUR 2023",
    "Total personnel costs th EUR 2022",
    "Added Value th EUR Last avail. yr",
    "Added Value th EUR 2024",
    "Added Value th EUR 2023",
    "Added Value th EUR 2022",
    "Turnover per employee EUR Last avail. yr",
    "Turnover per employee EUR 2024",
    "Turnover per employee EUR 2023",
    "Turnover per employee EUR 2022",
    "Added value per employee EUR Last avail. yr",
    "Added value per employee EUR 2024",
    "Added value per employee EUR 2023",
    "Added value per employee EUR 2022",
    "Staff Costs per employee EUR Last avail. yr",
    "Staff Costs per employee EUR 2024",
    "Staff Costs per employee EUR 2023",
    "Staff Costs per employee EUR 2022",
    "Turnover/Staff Costs Last avail. yr",
    "Turnover/Staff Costs 2024",
    "Turnover/Staff Costs 2023",
    "Turnover/Staff Costs 2022",
    "Number of directors & managers",
    "Number of current directors & managers",
    "Number of previous directors & managers",
    "Mergers & acquisitions deals",
    "No of companies in corporate group",
    "BvD independence indicator",
]

# 4. RELATIONAL CAPITAL: customers, suppliers, market reach, brand, network
relational_columns = [
    "Strategic alliances",
    "Main customers",
    "Import/export",
    "Membership of a network",
    "Subsidiary companies th EUR Last avail. yr",
    "Subsidiary companies th EUR 2024",
    "Subsidiary companies th EUR 2023",
    "Subsidiary companies th EUR 2022",
    "Associated companies th EUR Last avail. yr",
    "Associated companies th EUR 2024",
    "Associated companies th EUR 2023",
    "Associated companies th EUR 2022",
    "Main foreign countries or regions",
    "Main brand names",
    "Durata media dei crediti al lordo IVA (days) Last avail. yr",
    "Durata media dei crediti al lordo IVA (days) 2024",
    "Durata media dei crediti al lordo IVA (days) 2023",
    "Durata media dei crediti al lordo IVA (days) 2022",
    "Durata media dei debiti al lordo IVA (days) Last avail. yr",
    "Durata media dei debiti al lordo IVA (days) 2024",
    "Durata media dei debiti al lordo IVA (days) 2023",
    "Durata media dei debiti al lordo IVA (days) 2022",
]


# =============================================================================
# HELPERS
# =============================================================================
def _normalize(name: str) -> str:
    """Whitespace-insensitive form of a column name.

    AIDA exports often embed newlines inside header text
    ('Revenues from sales and services\\nth EUR\\nLast avail. yr').
    This collapses any whitespace run into a single space and lower-cases
    so the script matches whether your file uses newlines or spaces.
    """
    return re.sub(r"\s+", " ", str(name)).strip().lower()


def get_existing_columns(df: pd.DataFrame,
                         requested_columns: list[str],
                         group_name: str) -> list[str]:
    """Return the *actual* DataFrame column names matching the requested list.

    - Whitespace-insensitive matching (handles \\n vs space).
    - Preserves request order.
    - Prints a clear warning for any column not found, never crashes.
    """
    actual_lookup = {_normalize(c): c for c in df.columns}
    found, missing = [], []
    for req in requested_columns:
        hit = actual_lookup.get(_normalize(req))
        if hit is not None:
            found.append(hit)
        else:
            missing.append(req)

    print(f"  • {group_name:<32s} requested={len(requested_columns):>3d}  "
          f"found={len(found):>3d}  missing={len(missing):>3d}")
    if missing:
        print(f"    ⚠️  Missing in input file:")
        for m in missing:
            print(f"        - {m}")
    return found


def _populate_sheet(ws, df: pd.DataFrame) -> None:
    """Fill an openpyxl worksheet with a DataFrame and apply formatting in
    a single pass (no write-reload cycle, much faster for large files)."""
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill = PatternFill("solid", start_color="1F4E78")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---- Header row ------------------------------------------------------
    for j, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # ---- Data rows (bulk via numpy values for speed) ---------------------
    values = df.values
    n_rows, n_cols = values.shape
    for i in range(n_rows):
        excel_row = i + 2
        row = values[i]
        for j in range(n_cols):
            v = row[j]
            if v is None:
                continue
            # NaN check without pd.isna() overhead in tight loop
            if isinstance(v, float) and v != v:
                continue
            ws.cell(row=excel_row, column=j + 1, value=v)

    # ---- Layout ----------------------------------------------------------
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 45
    if n_rows > 0:
        ws.auto_filter.ref = ws.dimensions

    # ---- Column widths (header + small sample, capped) -------------------
    sample_n = min(n_rows, 200)
    for j, col_name in enumerate(df.columns, start=1):
        header_w = max((len(line) for line in str(col_name).split("\n")), default=10)
        data_w = 0
        if sample_n:
            try:
                m = df[col_name].head(sample_n).astype(str).str.len().max()
                if pd.notna(m):
                    data_w = int(m)
            except Exception:
                data_w = 0
        width = min(max(header_w, data_w) + 2, 50)
        ws.column_dimensions[get_column_letter(j)].width = max(width, 12)


def _write_single_sheet(df: pd.DataFrame, path: Path, sheet_name: str = "Sheet1") -> None:
    """Write a single-sheet xlsx, built directly with openpyxl."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    _populate_sheet(ws, df)
    wb.save(path)


def _write_multi_sheet(sheets: dict, path: Path) -> None:
    """Write a multi-sheet workbook, built directly with openpyxl."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    for name, df in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        _populate_sheet(ws, df)
    wb.save(path)


# =============================================================================
# MAIN
# =============================================================================
def main() -> int:
    # Resolve input file
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        user_input = input(
            f"Input file '{INPUT_FILE}' not found. "
            "Enter path to your AIDA export (or press Enter to abort): "
        ).strip().strip('"').strip("'")
        if not user_input:
            print("No input file provided. Aborting.")
            return 1
        input_path = Path(user_input)
        if not input_path.exists():
            print(f"File not found: {input_path}")
            return 1

    output_dir = Path(OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Reading: {input_path}")
    df = pd.read_excel(input_path, sheet_name=SHEET_NAME)
    print(f"Loaded {len(df):,} rows × {len(df.columns)} columns\n")

    # Resolve which columns actually exist in the file (whitespace-tolerant)
    print("Resolving column groups against input file:")
    ctx   = get_existing_columns(df, context_columns,            "Context")
    fin   = get_existing_columns(df, financial_columns,          "Financial Capital")
    tech  = get_existing_columns(df, technological_columns,      "Technological Capital")
    human = get_existing_columns(df, human_organisational_columns, "Human & Organisational")
    rel   = get_existing_columns(df, relational_columns,         "Relational Capital")

    # Build per-capital DataFrames (context first, then capital-specific)
    # Preserve original column names exactly as they appear in the input file.
    def _build(cap_cols: list[str]) -> pd.DataFrame:
        # context first, then capital cols, dropping any duplicates while preserving order
        seen, ordered = set(), []
        for c in ctx + cap_cols:
            if c not in seen:
                seen.add(c)
                ordered.append(c)
        return df[ordered].copy()

    context_df          = df[ctx].copy() if ctx else df.iloc[:, :0].copy()
    financial_df        = _build(fin)
    technological_df    = _build(tech)
    human_org_df        = _build(human)
    relational_df       = _build(rel)

    # ---- Write individual files ------------------------------------------
    print("\nWriting output files...")
    files = [
        (output_dir / "context_columns.xlsx",            context_df,       "Context"),
        (output_dir / "financial_capital.xlsx",          financial_df,     "Financial Capital"),
        (output_dir / "technological_capital.xlsx",      technological_df, "Technological Capital"),
        (output_dir / "human_organisational_capital.xlsx", human_org_df,   "Human & Organisational Capital"),
        (output_dir / "relational_capital.xlsx",         relational_df,    "Relational Capital"),
    ]
    written: list[Path] = []
    for path, frame, sheet in files:
        _write_single_sheet(frame, path, sheet)
        written.append(path)
        print(f"  ✓ {path}  ({frame.shape[0]:,} rows × {frame.shape[1]} cols)")

    # ---- Combined workbook ------------------------------------------------
    combined_path = output_dir / "all_capitals_clean_split.xlsx"
    _write_multi_sheet({
        "Context":                          context_df,
        "Financial Capital":                financial_df,
        "Technological Capital":            technological_df,
        "Human & Organisational Capital":   human_org_df,
        "Relational Capital":               relational_df,
    }, combined_path)
    written.append(combined_path)
    print(f"  ✓ {combined_path}  (5 sheets)")

    # ---- Final summary ----------------------------------------------------
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Input file              : {input_path}")
    print(f"Rows in input           : {len(df):,}")
    print(f"Total columns in input  : {len(df.columns)}")
    print(f"Context cols found      : {len(ctx):>3d} / {len(context_columns)}")
    print(f"Financial cols found    : {len(fin):>3d} / {len(financial_columns)}")
    print(f"Technological cols found: {len(tech):>3d} / {len(technological_columns)}")
    print(f"Human & Org cols found  : {len(human):>3d} / {len(human_organisational_columns)}")
    print(f"Relational cols found   : {len(rel):>3d} / {len(relational_columns)}")
    print(f"\nFiles written ({len(written)}):")
    for p in written:
        print(f"   - {p}")
    print()
    return 0


if __name__ == "__main__":
    sys.exit(main())
